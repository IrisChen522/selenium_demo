# coding:utf-8
import os
from datetime import time, datetime
import xlrd
from openpyxl import load_workbook
from xlrd import xldate_as_tuple, xldate_as_datetime

from config import global_param
PATH = lambda p: os.path.abspath(
    os.path.join(os.path.dirname(__file__), p)
)


class ExcelUtils:
    """对excel处理操作类"""
    # 当前目录上一层
    p = PATH('..')

    def __init__(self, book_name, sheet_name):
        book_path = os.path.join(self.p, "data\\"+global_param.language+"\\"+book_name)
        self.sheet = sheet_name
        self.wb = load_workbook(book_path)
        self.table = xlrd.open_workbook(book_path).sheet_by_name(sheet_name)

    def load_excel_by_path(self, file_path):
        return load_workbook(file_path)

    def get_cell_value_by_sheet_name_and_index(self,wb,sheet_name,col_index, row_index):
        ws = wb[sheet_name]
        return ws[col_index+row_index].value

    def get_cell_value(self, col_index, row_index):
        ws = self.wb[self.sheet]
        return ws[col_index+row_index].value

    def dict_data(self):
        nrows = self.table.nrows
        nclos = self.table.ncols

        if nrows <= 1:
            print("总行数小于1")
        else:
            r = []
            j = 1
            for i in list(range(nrows - 1)):
                s = {}
                # 从第二行取对应values值
                # s['rowNum'] = i+2
                values = self.table.row_values(j)
                for x in list(range(nclos)):
                    ctype = self.table.cell_type(j,x)  # 表格的数据类型
                    v = values[x]
                    if ctype == 2 and v % 1 == 0:  # 如果是整形
                        v = int(v)
                    elif ctype == 3:
                        # 转成datetime对象
                        if 0.0 <= v < 1.0:
                            v = xldate_as_datetime(v, 0).strftime('%H:%M')
                        else:
                            date = datetime(*xldate_as_tuple(v, 0))
                            v = date.strftime('%Y-%m-%d %H:%M')
                    elif ctype == 4:
                        v = True if v == 1 else False
                    s[self.table.row_values(0)[x]] = v
                r.append(s)
                j += 1
            return r



